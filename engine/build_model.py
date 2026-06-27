#!/usr/bin/env python3
"""
build_model.py -- generate the multi-tab financial-plan workbook from config.
Canadian edition (RRSP / TFSA / RRIF / LIRA / FHSA / RESP; CPP + OAS).

Mirrors the architecture documented in docs/ARCHITECTURE.md:
  * The ASSUMPTIONS tab is the single source of truth. Every other tab links
    back to it with cross-sheet formulas rather than hardcoding values.
  * Color convention:
      green text  = cross-sheet link  (=Assumptions!Cxx)
      black text  = intra-sheet formula
      blue text   = hardcoded input
  * Tabs built here: Assumptions, Net Worth Snapshot, Income Streams,
    Year-by-Year Projections, Employer Concentration, Monte Carlo (summary),
    RRSP Meltdown (lifetime-tax optimizer), Action Plan.

Run:
  python3 engine/build_model.py                      # demo config
  RPT_CONFIG=config/config.json python3 engine/build_model.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config_loader import (  # noqa: E402
    load_config, investable_total, employer_stock_total, employer_concentration_pct,
    current_age,
)
import tax_ca  # noqa: E402
from i18n import t, lang_of  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

GREEN = Font(color="008000")          # cross-sheet link
BLACK = Font(color="000000")          # intra-sheet formula
BLUE = Font(color="0000FF")           # hardcoded input
BOLD = Font(bold=True)
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="305496")
TITLE = Font(bold=True, size=14, color="1F3864")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
RED = Font(color="CF222E")
GOODF = Font(color="1A7F37", bold=True)

# RRIF prescribed minimum withdrawal factors (post-2015 rules), applied to the
# Jan-1 balance. Source: canada.ca "Chart - Prescribed factors" (see
# docs/CANADA_RULES.md s7). This is regulation, not user config, so it is
# encoded here rather than in config.json.
RRIF_FACTORS = {
    71: .0528, 72: .0540, 73: .0553, 74: .0567, 75: .0582, 76: .0598,
    77: .0617, 78: .0636, 79: .0658, 80: .0682, 81: .0708, 82: .0738,
    83: .0771, 84: .0808, 85: .0851, 86: .0899, 87: .0955, 88: .1021,
    89: .1099, 90: .1192, 91: .1306, 92: .1449, 93: .1634, 94: .1879,
}


def rrif_min_factor(age):
    """Prescribed RRIF minimum factor for an age. 0 before the first mandatory
    withdrawal (year you turn 72, after converting by 71); 20% at 95+."""
    if age < 72:
        return 0.0
    if age >= 95:
        return 0.20
    return RRIF_FACTORS.get(age, 0.0)


def _title(ws, text):
    ws["A1"] = text
    ws["A1"].font = TITLE
    ws.append([])


def _header_row(ws, cells, row=None):
    for col, val in enumerate(cells, start=1):
        c = ws.cell(row=row or ws.max_row, column=col, value=val)
        c.font = HDR
        c.fill = HDR_FILL
        c.border = BORDER


def _acct_label(key, lang="en"):
    """Localized human label for an account/holding/real-estate key.

    Looks the key up in the locale files (acct.<key>); falls back to a title-cased
    English derivation (with acronyms fixed) for any key not in the table."""
    label = t(f"acct.{key}", lang)
    if label != f"acct.{key}":
        return label
    pretty = key.replace("_", " ").title()
    for word, repl in (("Rrsp", "RRSP"), ("Tfsa", "TFSA"), ("Rrif", "RRIF"),
                       ("Lira", "LIRA"), ("Fhsa", "FHSA"), ("Resp", "RESP"),
                       ("Gics", "GICs")):
        pretty = pretty.replace(word, repl)
    return pretty


def build_assumptions(wb, cfg):
    lang = lang_of(cfg)
    ws = wb.create_sheet("Assumptions")
    _title(ws, t("wb.assumptions.title", lang))
    a = cfg["assumptions"]
    inc = cfg["income"]
    gb = cfg["government_benefits"]
    members = cfg["household"]["members"]
    na, nb = members[0]["display_name"], members[1]["display_name"]

    def L(k):  # localized assumption label / note shorthand
        return t(k, lang)

    rows = [
        (L("col.key"), L("col.value"), L("col.notes")),
        (L("asm.ret_base"), a["portfolio_return_base"], L("asmn.mc_mu")),
        (L("asm.ret_cons"), a["portfolio_return_conservative"], ""),
        (L("asm.ret_opt"), a["portfolio_return_optimistic"], ""),
        (L("asm.inflation"), a["inflation_rate"], L("asmn.bracket_scaling")),
        (L("asm.province"), cfg["household"]["province"], L("asmn.prov_juris")),
        (L("asm.bpa_fed"), a["basic_personal_amount_federal"], L("asmn.nonref")),
        (L("asm.bpa_prov"), a.get("provincial_basic_personal_amount", 0), ""),
        (L("asm.oas_thresh"), a["oas_clawback_threshold"], L("asmn.meltdown_ceiling")),
        (L("asm.oas_full"), a.get("oas_full_clawback", 0), L("asmn.oas_full")),
        (L("asm.rrif_age"), a.get("rrif_conversion_age", 71), L("asmn.rrif_age")),
        (L("asm.capgains"), a.get("capital_gains_inclusion_rate", 0.5), L("asmn.capgains")),
        (L("asm.rrsp_limit"), a.get("rrsp_dollar_limit", 0), L("asmn.rrsp_limit")),
        (L("asm.tfsa_limit"), a.get("tfsa_annual_limit", 0), ""),
        (L("asm.spend"), a["retirement_spend_annual"], ""),
        (L("asm.eq"), a["target_equity_pct"], ""),
        (L("asm.bond"), a["target_bond_pct"], ""),
        (L("asm.bridge"), a["bridge_target"], L("asmn.bridge")),
        (L("asm.pension_m"), inc["pension_monthly_at_retirement"], L("asmn.db_pension")),
        (L("asm.cola"), inc["pension_cola"], ""),
        (L("asm.passive"), inc["passive_income_annual"], ""),
        (f"{na} {L('asm.s.salary')}", inc["spouse_a_salary"], ""),
        (f"{na} {L('asm.s.bonus')}", inc["spouse_a_bonus_pct"], ""),
        (f"{na} {L('asm.s.rsu')}", inc["spouse_a_rsu_annual"], ""),
        (f"{nb} {L('asm.s.income')}", inc["spouse_b_annual"], ""),
        (f"{na} {L('asm.s.ret_age')}", members[0]["retirement_age"], ""),
        (f"{nb} {L('asm.s.ret_age')}", members[1]["retirement_age"], ""),
        (f"{na} {L('asm.s.cpp_age')}", members[0]["cpp_claim_age"], L("asmn.cpp_range")),
        (f"{na} {L('asm.s.oas_age')}", members[0]["oas_claim_age"], L("asmn.oas_range")),
        (f"{nb} {L('asm.s.cpp_age')}", members[1]["cpp_claim_age"], ""),
        (f"{nb} {L('asm.s.oas_age')}", members[1]["oas_claim_age"], ""),
        (f"{na} {L('asm.s.cpp_m')}", gb["spouse_a_cpp_monthly"], ""),
        (f"{na} {L('asm.s.oas_m')}", gb["spouse_a_oas_monthly"], ""),
        (f"{nb} {L('asm.s.cpp_m')}", gb["spouse_b_cpp_monthly"], ""),
        (f"{nb} {L('asm.s.oas_m')}", gb["spouse_b_oas_monthly"], ""),
    ]
    start = ws.max_row + 1
    for i, r in enumerate(rows):
        ws.append(r)
        if i == 0:
            _header_row(ws, r)
        else:
            ws.cell(row=ws.max_row, column=2).font = BLUE  # hardcoded inputs
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 44
    cfg["_assum_rows"] = {name: start + idx for idx, (name, *_ ) in enumerate(rows)}
    return ws


def build_net_worth(wb, cfg):
    lang = lang_of(cfg)
    ws = wb.create_sheet("Net Worth Snapshot")
    _title(ws, t("wb.networth.title", lang))
    ws.append([t("col.account", lang), t("col.balance", lang)])
    _header_row(ws, [t("col.account", lang), t("col.balance", lang)])
    first_data = ws.max_row + 1
    for k, v in cfg["accounts"].items():
        ws.append([_acct_label(k, lang), v])
        ws.cell(row=ws.max_row, column=2).font = BLUE
    for k, v in cfg["real_estate"].items():
        if v:
            ws.append([_acct_label(k, lang), v])
            ws.cell(row=ws.max_row, column=2).font = BLUE
    last_data = ws.max_row
    ws.append([t("nw.total", lang), f"=SUM(B{first_data}:B{last_data})"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.cell(row=ws.max_row, column=2).font = BLACK
    inv = investable_total(cfg)
    ws.append([t("nw.investable", lang), inv])
    ws.cell(row=ws.max_row, column=2).font = BLUE
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    return ws


def build_concentration(wb, cfg):
    lang = lang_of(cfg)
    es = cfg["employer_stock"]
    ws = wb.create_sheet("Employer Concentration")
    _title(ws, t("wb.conc.title", lang, emp=es["employer_name"], tk=es["ticker"]))
    ws.append([t("col.sleeve", lang), t("col.value", lang)])
    _header_row(ws, [t("col.sleeve", lang), t("col.value", lang)])
    for k, v in es["holdings"].items():
        ws.append([_acct_label(k, lang), v])
        ws.cell(row=ws.max_row, column=2).font = BLUE
    ws.append([t("conc.total_exposure", lang), employer_stock_total(cfg)])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append([t("conc.investable_total", lang), investable_total(cfg)])
    ws.append([t("conc.concentration_pct", lang), employer_concentration_pct(cfg)])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append([t("conc.watch_threshold", lang), es["watch_threshold_pct"]])
    ws.append([t("conc.trim_threshold", lang), es["trim_threshold_pct"]])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    return ws


def build_year_by_year(wb, cfg):
    lang = lang_of(cfg)
    ws = wb.create_sheet("Year-by-Year Projections")
    _title(ws, t("wb.yby.title", lang))
    members = cfg["household"]["members"]
    a = cfg["assumptions"]
    spend0 = a["retirement_spend_annual"]
    infl = a["inflation_rate"]
    ret = a["portfolio_return_base"]
    pension = cfg["income"]["pension_monthly_at_retirement"] * 12
    passive = cfg["income"]["passive_income_annual"]
    gb = cfg["government_benefits"]
    cpp_a = gb["spouse_a_cpp_monthly"] * 12
    oas_a = gb["spouse_a_oas_monthly"] * 12
    cpp_b = gb["spouse_b_cpp_monthly"] * 12
    oas_b = gb["spouse_b_oas_monthly"] * 12

    headers = [t("yby.year", lang), t("yby.age_a", lang), t("yby.age_b", lang),
               t("yby.spend", lang), t("yby.pension", lang), t("yby.passive", lang),
               t("yby.cpp_oas", lang), t("yby.draw", lang), t("yby.eoy", lang)]
    ws.append(headers)
    _header_row(ws, headers)

    import datetime
    base_year = datetime.date.today().year
    a_age0 = current_age(cfg, "spouse_a")
    b_age0 = current_age(cfg, "spouse_b")
    a_ret_age = members[0]["retirement_age"]
    a_cpp_age, a_oas_age = members[0]["cpp_claim_age"], members[0]["oas_claim_age"]
    b_cpp_age, b_oas_age = members[1]["cpp_claim_age"], members[1]["oas_claim_age"]
    portfolio = investable_total(cfg)

    for n in range(0, 36):
        year = base_year + n
        a_age = a_age0 + n
        b_age = b_age0 + n
        retired = a_age >= a_ret_age
        spend = spend0 * ((1 + infl) ** n) if retired else 0
        pen = pension * ((1 + cfg["income"]["pension_cola"]) ** max(0, a_age - a_ret_age)) if retired else 0
        pas = passive if retired else 0
        govt = ((cpp_a if a_age >= a_cpp_age else 0) + (oas_a if a_age >= a_oas_age else 0)
                + (cpp_b if b_age >= b_cpp_age else 0) + (oas_b if b_age >= b_oas_age else 0))
        portfolio *= (1 + ret)
        draw = 0
        if retired:
            need = spend - pen - pas - govt
            draw = max(0, need)
            portfolio -= draw
        ws.append([year, a_age, b_age, round(spend), round(pen), round(pas),
                   round(govt), round(draw), round(portfolio)])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    return ws


def build_monte_carlo(wb, cfg):
    lang = lang_of(cfg)
    ws = wb.create_sheet("Monte Carlo")
    _title(ws, t("wb.mc.title", lang))
    hdr = [t("mc.col.scenario", lang), t("mc.col.return_mu", lang),
           t("mc.col.success_rate", lang), t("mc.col.note", lang)]
    ws.append(hdr)
    _header_row(ws, hdr)
    a = cfg["assumptions"]
    for key, mu in [("conservative", a["portfolio_return_conservative"]),
                    ("base", a["portfolio_return_base"]),
                    ("optimistic", a["portfolio_return_optimistic"])]:
        ws.append([t(f"scenario.{key}", lang), mu,
                   t("mc.cell.pending", lang), t("mc.cell.pending_note", lang)])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["D"].width = 40
    return ws


# ===========================================================================
# RRSP Meltdown -- lifetime-tax optimizer
# ===========================================================================
def _simulate_meltdown(cfg, strategy, target=None):
    """Simulate the household's registered drawdown to a planning horizon and
    return total lifetime tax + a year-by-year schedule.

    strategy:
      "none"     -- take only the forced RRIF minimum (the do-nothing baseline)
      "clawback" -- fill household income up to (n_spouses x OAS clawback line)
      "optimal"  -- fill each retired spouse's taxable income up to `target`

    Lifetime tax = sum of annual (income tax + OAS clawback) for both spouses,
    PLUS the terminal deemed-disposition tax on the RRSP/RRIF still standing at
    the horizon (taxed to a single surviving filer -- the reason melting early
    can win). All taxes are discounted to PRESENT VALUE at the inflation rate, so
    the optimizer trades the cost of paying tax now against deferring it -- this
    is what produces an interior optimum rather than "drain as fast as possible."
    See docs/CANADA_RULES.md for the modelling assumptions.
    """
    import datetime
    a = cfg["assumptions"]
    m = cfg["household"]["members"]
    gb, inc, acct = cfg["government_benefits"], cfg["income"], cfg["accounts"]
    prov = cfg["household"]["province"]

    infl, ret = a["inflation_rate"], a["portfolio_return_base"]
    thr0 = a["oas_clawback_threshold"]
    spend0 = a["retirement_spend_annual"]
    base_year = datetime.date.today().year

    a_age0, b_age0 = current_age(cfg, "spouse_a"), current_age(cfg, "spouse_b")
    a_ret, b_ret = m[0]["retirement_age"], m[1]["retirement_age"]
    a_cpp, a_oas = m[0]["cpp_claim_age"], m[0]["oas_claim_age"]
    b_cpp, b_oas = m[1]["cpp_claim_age"], m[1]["oas_claim_age"]

    rrsp = float(acct.get("spouse_a_rrsp", 0) + acct.get("spouse_b_rrsp", 0))
    buffer = float(acct.get("spouse_a_non_registered", 0) + acct.get("spouse_b_non_registered", 0)
                   + acct.get("joint_non_registered", 0) + acct.get("cash_and_gics", 0))
    tfsa = float(acct.get("spouse_a_tfsa", 0) + acct.get("spouse_b_tfsa", 0))

    pension_m, cola = inc["pension_monthly_at_retirement"], inc["pension_cola"]
    passive0, b_salary0 = inc["passive_income_annual"], inc["spouse_b_annual"]
    cpp_a0, oas_a0 = gb["spouse_a_cpp_monthly"] * 12, gb["spouse_a_oas_monthly"] * 12
    cpp_b0, oas_b0 = gb["spouse_b_cpp_monthly"] * 12, gb["spouse_b_oas_monthly"] * 12

    horizon = max(1, 90 - a_age0)            # project to spouse A age 90
    lifetime_tax, insolvent, schedule = 0.0, False, []
    last_survivor_base = 0.0

    for n in range(0, horizon + 1):
        year = base_year + n
        a_age, b_age = a_age0 + n, b_age0 + n
        rrsp *= (1 + ret); buffer *= (1 + ret); tfsa *= (1 + ret)
        if a_age < a_ret:
            continue  # still working -- no draws modelled before retirement

        idx = (1 + infl) ** n
        spend = spend0 * idx
        pension = pension_m * 12 * ((1 + cola) ** max(0, a_age - a_ret))
        passive = passive0 * idx
        b_work = b_salary0 * idx if b_age < b_ret else 0.0
        cpp_a = cpp_a0 * idx if a_age >= a_cpp else 0.0
        oas_a = oas_a0 * idx if a_age >= a_oas else 0.0
        cpp_b = cpp_b0 * idx if b_age >= b_cpp else 0.0
        oas_b = oas_b0 * idx if b_age >= b_oas else 0.0

        forced = min(rrsp * rrif_min_factor(a_age), rrsp)
        retirement_fixed = pension + passive + cpp_a + oas_a + cpp_b + oas_b + forced
        n_active = (1 if a_age >= a_ret else 0) + (1 if b_age >= b_ret else 0)

        if strategy == "none":
            voluntary = 0.0
        elif strategy == "clawback":
            ceiling = n_active * thr0 * idx
            voluntary = max(0.0, min(ceiling - (retirement_fixed + b_work), rrsp - forced))
        else:  # optimal -- level per-spouse taxable target
            ceiling = n_active * target
            voluntary = max(0.0, min(ceiling - (retirement_fixed + b_work), rrsp - forced))

        withdraw = forced + voluntary
        rrsp -= withdraw

        # ---- tax: per-spouse, equalizing retirement income once both are retired.
        # Threads age (age amount), eligible pension income (pension credit),
        # net family income (Quebec's family-tested credit), and the HSF base
        # (income excl. OAS/employment) into the tax engine. ----
        retire_income = retirement_fixed + voluntary
        reg_withdraw = forced + voluntary
        both_retired = (a_age >= a_ret) and (b_age >= b_ret)
        family_total = retire_income + b_work
        thr_n = thr0 * idx
        if both_retired:
            half = retire_income / 2.0
            oas_each = (oas_a + oas_b) / 2.0
            # eligible pension income: DB pension always; registered (RRIF)
            # withdrawals qualify only at 65+ (gate on the older spouse, A).
            elig_each = (pension + (reg_withdraw if a_age >= 65 else 0.0)) / 2.0
            hsf_each = max(0.0, half - oas_each)
            tax = (tax_ca.income_tax(half, prov, year, infl, age=a_age, pension_income=elig_each,
                                     family_net_income=retire_income, hsf_base=hsf_each)
                   + tax_ca.income_tax(half, prov, year, infl, age=b_age, pension_income=elig_each,
                                       family_net_income=retire_income, hsf_base=hsf_each))
            claw = 2 * tax_ca.oas_clawback(half, oas_each, thr_n)
        else:
            inc_a, inc_b = retire_income, b_work
            elig_a = pension + (reg_withdraw if a_age >= 65 else 0.0)
            tax = (tax_ca.income_tax(inc_a, prov, year, infl, age=a_age, pension_income=elig_a,
                                     family_net_income=family_total, hsf_base=max(0.0, inc_a - oas_a))
                   + tax_ca.income_tax(inc_b, prov, year, infl, age=b_age, pension_income=0.0,
                                       family_net_income=family_total, hsf_base=0.0))  # B = employment, no HSF
            claw = (tax_ca.oas_clawback(inc_a, oas_a, thr_n)
                    + tax_ca.oas_clawback(inc_b, oas_b, thr_n))
        last_survivor_base = pension + cpp_a + oas_a
        year_tax = tax + claw
        lifetime_tax += year_tax / ((1 + infl) ** n)   # present value, today's $

        # ---- fund spending; surplus registered cash sweeps into the TFSA ----
        cash_in = pension + passive + cpp_a + oas_a + cpp_b + oas_b + b_work + withdraw
        net_cash = cash_in - year_tax
        if net_cash >= spend:
            tfsa += (net_cash - spend)
        else:
            short = spend - net_cash
            take = min(buffer, short); buffer -= take; short -= take
            if short > 0:
                take = min(tfsa, short); tfsa -= take; short -= take
            if short > 1.0:
                insolvent = True

        schedule.append({
            "year": year, "a_age": a_age, "b_age": b_age,
            "taxable": retire_income, "forced": forced, "voluntary": voluntary,
            "tax": tax, "claw": claw, "rrsp": rrsp, "estate": rrsp + buffer + tfsa,
            "over": claw > 0,
        })

    # ---- terminal tax: RRSP left standing is deemed disposed at death (single filer) ----
    final_year = base_year + horizon
    final_age = a_age0 + horizon   # surviving spouse, single filer
    terminal_raw = (tax_ca.income_tax(last_survivor_base + rrsp, prov, final_year, infl,
                                      age=final_age, pension_income=last_survivor_base + rrsp)
                    - tax_ca.income_tax(last_survivor_base, prov, final_year, infl,
                                        age=final_age, pension_income=last_survivor_base))
    terminal_tax = terminal_raw / ((1 + infl) ** horizon)   # present value, today's $
    total_tax = lifetime_tax + terminal_tax
    estate = rrsp + buffer + tfsa
    return {
        "strategy": strategy, "target": target,
        "lifetime_tax": lifetime_tax, "terminal_tax": terminal_tax, "total_tax": total_tax,
        "rrsp_end": rrsp, "estate_end": estate, "insolvent": insolvent, "schedule": schedule,
    }


def _optimize_meltdown(cfg):
    """Grid-search the level per-spouse meltdown target that minimizes total
    lifetime tax (subject to the plan staying solvent)."""
    best = None
    for t in range(20000, 200001, 2500):   # wide enough that large RRSPs don't clip the optimum
        r = _simulate_meltdown(cfg, "optimal", target=float(t))
        if r["insolvent"]:
            continue
        if best is None or r["total_tax"] < best["total_tax"]:
            best = r
    if best is None:  # fall back to the lowest target if nothing is solvent
        best = _simulate_meltdown(cfg, "optimal", target=20000.0)
    return best


def build_meltdown(wb, cfg):
    """RRSP-meltdown / OAS-clawback tab -- lifetime-tax optimizer (objective #2).

    Searches for the level annual RRSP/RRIF withdrawal target (per spouse) that
    minimizes TOTAL lifetime tax -- in-life income tax + OAS clawback across both
    spouses, PLUS the terminal deemed-disposition tax on whatever RRSP is left at
    death. Compares three strategies and shows the winner's year-by-year plan.

    Tax engine: engine/tax_ca.py (federal + Ontario + Quebec; other provinces fall
    back to Ontario with a warning). Models ordinary income, the age + pension-income
    credits, the OAS clawback, Quebec's HSF contribution, pension/registered income
    equalization between spouses, and terminal RRSP tax. Does NOT yet model
    non-registered capital-gains tax, the dividend tax credit, or TFSA contribution
    limits. Illustrative, not advice.
    """
    lang = lang_of(cfg)
    ws = wb.create_sheet("RRSP Meltdown")
    _title(ws, t("wb.meltdown.title", lang))
    prov = cfg["household"]["province"]

    none = _simulate_meltdown(cfg, "none")
    claw = _simulate_meltdown(cfg, "clawback")
    best = _optimize_meltdown(cfg)

    ws.append([t("mlt.objective", lang, prov=prov)])
    ws.append([])

    # ---- strategy comparison ----
    cmp_headers = [t("mlt.col.strategy", lang), t("mlt.col.target", lang), t("mlt.col.inlife", lang),
                   t("mlt.col.terminal", lang), t("mlt.col.total", lang),
                   t("mlt.col.rrsp_horizon", lang), t("mlt.col.estate_horizon", lang)]
    ws.append(cmp_headers)
    _header_row(ws, cmp_headers)
    rows = [
        (t("mlt.strat.none", lang), t("mlt.target.na", lang), none),
        (t("mlt.strat.clawback", lang), t("mlt.target.clawback", lang), claw),
        (t("mlt.strat.optimal", lang), t("mlt.target.per_year", lang, v=f"{best['target']:,.0f}"), best),
    ]
    base_total = none["total_tax"]
    for label, tgt, r in rows:
        ws.append([label, tgt, round(r["lifetime_tax"]), round(r["terminal_tax"]),
                   round(r["total_tax"]), round(r["rrsp_end"]), round(r["estate_end"])])
        if r is best:
            for c in range(1, len(cmp_headers) + 1):
                ws.cell(row=ws.max_row, column=c).font = GOODF
    ws.append([])
    saved = base_total - best["total_tax"]
    ws.append([t("mlt.saved_label", lang), round(saved),
               t("mlt.saved_note", lang, pct=f"{100*saved/base_total:.0f}", amt=f"{best['target']:,.0f}")])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.cell(row=ws.max_row, column=2).font = GOODF
    ws.append([])

    # ---- optimal year-by-year schedule ----
    sch_headers = [t("mlt.sch.year", lang), t("mlt.sch.age_a", lang), t("mlt.sch.age_b", lang),
                   t("mlt.sch.taxable", lang), t("mlt.sch.rrif_min", lang), t("mlt.sch.melt", lang),
                   t("mlt.sch.income_tax", lang), t("mlt.sch.clawback", lang), t("mlt.sch.rrsp_bal", lang)]
    ws.append(sch_headers)
    _header_row(ws, sch_headers)
    for r in best["schedule"]:
        ws.append([r["year"], r["a_age"], r["b_age"], round(r["taxable"]),
                   round(r["forced"]), round(r["voluntary"]), round(r["tax"]),
                   round(r["claw"]), round(r["rrsp"])])
        if r["claw"] > 0:
            ws.cell(row=ws.max_row, column=8).font = RED

    widths = [8, 7, 7, 16, 14, 17, 13, 14, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def build_action_plan(wb, cfg):
    lang = lang_of(cfg)
    rec, opn = t("status.recurring", lang), t("status.open", lang)
    ws = wb.create_sheet("Action Plan")
    _title(ws, t("wb.action.title", lang))
    hdr = [t("col.priority", lang), t("col.item", lang), t("col.status", lang)]
    ws.append(hdr)
    _header_row(ws, hdr)
    items = [
        (1, t("action.concentration", lang), rec),
        (2, t("action.tfsa", lang), rec),
        (3, t("action.meltdown", lang), opn),
        (4, t("action.rrif", lang), opn),
        (5, t("action.splitting", lang), opn),
        (6, t("action.beneficiary", lang), opn),
        (7, t("action.survivor", lang), opn),
        (8, t("action.rebalance", lang), rec),
    ]
    for p, it, st in items:
        ws.append([p, it, st])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 14
    return ws


def build_income_note(wb, cfg):
    lang = lang_of(cfg)
    ws = wb.create_sheet("Income Streams")
    _title(ws, t("wb.income.title", lang))
    hdr = [t("col.stream", lang), t("col.annual", lang), t("col.notes", lang)]
    ws.append(hdr)
    _header_row(ws, hdr)
    inc = cfg["income"]
    gb = cfg["government_benefits"]
    m = cfg["household"]["members"]
    na, nb = m[0]["display_name"], m[1]["display_name"]
    rows = [
        (f"{na} {t('asm.s.salary', lang)}", inc["spouse_a_salary"], ""),
        (f"{na} {t('inc.s.bonus', lang)}", round(inc["spouse_a_salary"] * inc["spouse_a_bonus_pct"]), ""),
        (f"{na} {t('inc.s.rsu', lang)}", inc["spouse_a_rsu_annual"], t("inc.rsu_note", lang)),
        (f"{nb} {t('inc.s.income', lang)}", inc["spouse_b_annual"], ""),
        (t("inc.pension", lang), inc["pension_monthly_at_retirement"] * 12, t("asmn.db_pension", lang)),
        (f"{na} {t('inc.s.cpp', lang)}", gb["spouse_a_cpp_monthly"] * 12, t("inc.claim_age", lang, age=m[0]["cpp_claim_age"])),
        (f"{na} {t('inc.s.oas', lang)}", gb["spouse_a_oas_monthly"] * 12, t("inc.claim_age_claw", lang, age=m[0]["oas_claim_age"])),
        (f"{nb} {t('inc.s.cpp', lang)}", gb["spouse_b_cpp_monthly"] * 12, t("inc.claim_age", lang, age=m[1]["cpp_claim_age"])),
        (f"{nb} {t('inc.s.oas', lang)}", gb["spouse_b_oas_monthly"] * 12, t("inc.claim_age_claw", lang, age=m[1]["oas_claim_age"])),
        (t("inc.passive", lang), inc["passive_income_annual"], ""),
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 44
    return ws


def build(cfg):
    wb = Workbook()
    wb.remove(wb.active)
    build_assumptions(wb, cfg)
    build_net_worth(wb, cfg)
    build_concentration(wb, cfg)
    build_income_note(wb, cfg)
    build_year_by_year(wb, cfg)
    build_monte_carlo(wb, cfg)
    build_meltdown(wb, cfg)
    build_action_plan(wb, cfg)
    return wb


def main():
    cfg, path = load_config()
    out = ROOT / cfg["paths"]["model_xlsx"]
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build(cfg)
    wb.save(out)
    tag = "DEMO" if cfg.get("_is_demo") else "LIVE"
    print(f"[{tag}] Built model from {path.name}")
    print(f"  -> {out}")
    print(f"  Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
