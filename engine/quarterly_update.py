#!/usr/bin/env python3
"""
quarterly_update.py -- single entry point for a quarterly refresh.

Pipeline (mirrors docs/QUARTERLY_WORKFLOW.md):
  1. Load config + the quarterly input JSON (new account balances).
  2. Apply balances to config (real-estate carries forward when null).
  3. Rebuild the workbook (build_model.build).
  4. Run a Monte Carlo simulation (3 scenarios) on the investable total.
  5. Write a Monte Carlo summary JSON and stamp it into the workbook.
  6. Refresh the HTML dashboard from the latest state.

This does NOT mutate config.json on disk -- balances are applied to an
in-memory copy for the run, then you decide what to persist. Keeps the source
of truth explicit.

Run:
  python3 engine/quarterly_update.py --input templates/quarterly_input_TEMPLATE.json
  RPT_CONFIG=config/config.json python3 engine/quarterly_update.py --input quarterly_input_Q3_2026.json
"""

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config_loader import load_config, investable_total  # noqa: E402
import build_model  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent

try:
    import numpy as np
except ImportError:
    np = None


def apply_input(cfg, input_path):
    """Overlay account balances from the quarterly input JSON onto the config."""
    if not input_path:
        return cfg, "no input file (using current config balances)"
    with open(input_path, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    applied = 0
    for k, v in (data.get("accounts") or {}).items():
        if v is not None and k in cfg["accounts"]:
            cfg["accounts"][k] = v
            applied += 1
    for k, v in (data.get("real_estate") or {}).items():
        if v is not None and k in cfg["real_estate"]:
            cfg["real_estate"][k] = v  # null = carry forward (skip)
    es = data.get("employer_stock_holdings") or {}
    for k, v in es.items():
        if v is not None and k in cfg["employer_stock"]["holdings"]:
            cfg["employer_stock"]["holdings"][k] = v
    return cfg, f"applied {applied} account balance(s) from {Path(input_path).name}"


def monte_carlo(cfg, n_sims=10000, years=30, seed=42):
    """Return per-scenario success rates. Success = portfolio > 0 at horizon."""
    if np is None:
        return {"error": "numpy not installed -- pip install numpy"}
    rng = np.random.default_rng(seed)
    a = cfg["assumptions"]
    start = investable_total(cfg)
    spend = a["retirement_spend_annual"]
    infl = a["inflation_rate"]
    sigma = 0.12  # annual volatility assumption
    # income floor offsets the draw (pension + passive + CPP/OAS approximated flat).
    # CPP/OAS are time-phased in the year-by-year tab; here they are averaged in at
    # ~half weight as a steady-state floor for the success-rate approximation.
    gb = cfg["government_benefits"]
    govt_annual = (gb["spouse_a_cpp_monthly"] + gb["spouse_a_oas_monthly"]
                   + gb["spouse_b_cpp_monthly"] + gb["spouse_b_oas_monthly"]) * 12
    floor = (cfg["income"]["pension_monthly_at_retirement"] * 12
             + cfg["income"]["passive_income_annual"]
             + govt_annual * 0.5)
    out = {}
    for label, mu in [("conservative", a["portfolio_return_conservative"]),
                      ("base", a["portfolio_return_base"]),
                      ("optimistic", a["portfolio_return_optimistic"])]:
        bal = np.full(n_sims, float(start))
        alive = np.ones(n_sims, dtype=bool)
        for y in range(years):
            r = rng.normal(mu, sigma, n_sims)
            bal = bal * (1 + r)
            draw = max(0.0, spend * ((1 + infl) ** y) - floor)
            bal = bal - draw
            alive &= bal > 0
        out[label] = {
            "mu": mu,
            "success_rate": round(float(alive.mean()) * 100, 1),
            "median_end_balance": round(float(np.median(bal))),
        }
    out["_params"] = {"n_sims": n_sims, "years": years, "start": start, "annual_spend": spend}
    return out


def stamp_mc(xlsx_path, mc):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    if "Monte Carlo" not in wb.sheetnames:
        return
    ws = wb["Monte Carlo"]
    label_to_row = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v in ("Conservative", "Base", "Optimistic"):
            label_to_row[v.lower()] = row[0].row
    for label, row in label_to_row.items():
        if label in mc:
            ws.cell(row=row, column=3, value=f"{mc[label]['success_rate']}%")
            ws.cell(row=row, column=4,
                    value=f"median end ${mc[label]['median_end_balance']:,}")
    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description="Quarterly refresh pipeline")
    ap.add_argument("--input", help="Quarterly input JSON with new balances")
    ap.add_argument("--sims", type=int, default=10000)
    ap.add_argument("--no-dashboard", action="store_true")
    args = ap.parse_args()

    cfg, path = load_config()
    tag = "DEMO" if cfg.get("_is_demo") else "LIVE"
    print(f"[{tag}] Quarterly update using {path.name}")

    cfg, msg = apply_input(cfg, args.input)
    print(f"  Step 1-2: {msg}")

    out = ROOT / cfg["paths"]["model_xlsx"]
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_model.build(cfg)
    wb.save(out)
    print(f"  Step 3: rebuilt workbook -> {out.name}  (investable ${investable_total(cfg):,.0f})")

    mc = monte_carlo(cfg, n_sims=args.sims)
    mc_path = ROOT / "model" / "mc_summary.json"
    with open(mc_path, "w", encoding="utf-8") as fh:
        json.dump(mc, fh, indent=2)
    if "error" not in mc:
        stamp_mc(out, mc)
        print(f"  Step 4-5: Monte Carlo -> base success {mc['base']['success_rate']}% "
              f"(cons {mc['conservative']['success_rate']}% / opt {mc['optimistic']['success_rate']}%)")
    else:
        print(f"  Step 4-5: {mc['error']}")

    if not args.no_dashboard:
        try:
            import refresh_dashboard
            refresh_dashboard.refresh(cfg, mc)
            print("  Step 6: dashboard refreshed.")
        except Exception as exc:  # noqa: BLE001
            print(f"  Step 6: dashboard refresh skipped ({exc.__class__.__name__}: {exc})")

    print("  Done.")


if __name__ == "__main__":
    main()
