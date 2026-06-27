"""Integration tests: both demos build the full workbook, render the dashboard,
and run the Monte Carlo pipeline without error."""
import pathlib

import openpyxl
import pytest
import config_loader as cl
import build_model as bm
import refresh_dashboard as rd
import quarterly_update as qu

ROOT = pathlib.Path(__file__).resolve().parent.parent
EXPECTED_SHEETS = {
    "Assumptions", "Net Worth Snapshot", "Employer Concentration", "Income Streams",
    "Year-by-Year Projections", "Monte Carlo", "RRSP Meltdown", "Action Plan",
}
DEMOS = ["tremblay_config.json", "gagnon_config.json"]


def _cfg(name):
    cfg, _ = cl.load_config(str(ROOT / "config" / "examples" / name))
    return cfg


@pytest.mark.parametrize("name", DEMOS)
def test_build_produces_all_tabs(name):
    wb = bm.build(_cfg(name))
    assert set(wb.sheetnames) == EXPECTED_SHEETS


def test_built_workbook_round_trips(tmp_path):
    wb = bm.build(_cfg("tremblay_config.json"))
    out = tmp_path / "plan.xlsx"
    wb.save(out)
    reloaded = openpyxl.load_workbook(out)
    assert "RRSP Meltdown" in reloaded.sheetnames
    # the meltdown tab should carry a numeric "TOTAL lifetime tax" comparison
    ws = reloaded["RRSP Meltdown"]
    cells = [c.value for row in ws.iter_rows() for c in row]
    assert any(isinstance(v, str) and "lifetime-tax optimal" in v.lower() for v in cells)


@pytest.mark.parametrize("name", DEMOS)
def test_dashboard_renders_with_panels(name):
    # Language-neutral structural check (EN/FR text is covered by dedicated tests).
    cfg = _cfg(name)
    doc = rd.render(cfg, qu.monte_carlo(cfg, n_sims=300))
    assert "<!DOCTYPE html>" in doc
    assert cfg["household"]["name"] in doc
    assert 'class="kpis"' in doc            # KPI tiles present
    assert doc.count('class="panel"') >= 2  # concentration + Monte Carlo panels


def test_dashboard_localized_french():
    # The Quebec demo is configured language=fr -> French dashboard.
    cfg = _cfg("gagnon_config.json")
    doc = rd.render(cfg, qu.monte_carlo(cfg, n_sims=200))
    assert 'lang="fr"' in doc
    assert "Tableau de bord de retraite" in doc
    assert "Valeur nette totale" in doc
    assert "RPC/SV" in doc                       # CPP/OAS localized
    assert "Gagnon" in doc and "Tremblay" not in doc   # banner uses the right household


def test_dashboard_default_english():
    cfg = _cfg("tremblay_config.json")
    doc = rd.render(cfg, qu.monte_carlo(cfg, n_sims=200))
    assert 'lang="en"' in doc
    assert "Retirement Dashboard" in doc and "Total net worth" in doc


def test_workbook_localized_french():
    cfg = _cfg("gagnon_config.json")           # configured language=fr
    wb = bm.build(cfg)
    # sheet/tab names stay English (stable identifiers used by stamp_mc)
    assert EXPECTED_SHEETS.issubset(set(wb.sheetnames))
    assert wb["Assumptions"]["A1"].value.startswith("Hypothèses")
    assert wb["Net Worth Snapshot"]["A4"].value == "REER conjoint A"   # spouse_a_rrsp localized
    assert wb["RRSP Meltdown"]["A1"].value.startswith("Fonte du REER")


def test_workbook_default_english():
    cfg = _cfg("tremblay_config.json")
    wb = bm.build(cfg)
    assert wb["Assumptions"]["A1"].value.startswith("Assumptions")
    assert wb["Net Worth Snapshot"]["A4"].value == "Spouse A RRSP"


def test_mc_stamping_is_locale_aware(tmp_path):
    # The French workbook has localized scenario rows; stamp_mc must still find
    # and populate them (it keys off every locale's label, not just English).
    import openpyxl
    cfg = _cfg("gagnon_config.json")
    out = tmp_path / "fr.xlsx"
    bm.build(cfg).save(out)
    qu.stamp_mc(str(out), qu.monte_carlo(cfg, n_sims=200))
    ws = openpyxl.load_workbook(out)["Monte Carlo"]
    stamped = {r[0].value: r[2].value for r in ws.iter_rows(min_row=4, max_row=6)}
    assert "Prudent" in stamped and stamped["Prudent"].endswith("%")


def test_dashboard_renders_without_monte_carlo():
    # Should degrade gracefully when no MC summary is supplied.
    html = rd.render(_cfg("tremblay_config.json"), mc=None)
    assert "<!DOCTYPE html>" in html


def test_quarterly_apply_input_overlays_balances():
    cfg = _cfg("tremblay_config.json")
    before = cfg["accounts"]["spouse_a_rrsp"]
    cfg2, msg = qu.apply_input(cfg, None)   # no input file -> unchanged
    assert cfg2["accounts"]["spouse_a_rrsp"] == before
